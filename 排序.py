import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

# 设置页面标题
st.title("📊 Excel数据处理工具 (多条件排序版)")
st.write("支持按学院精确排序或按时间列排序")

# 上传Excel文件
excel_file = st.file_uploader("选择Excel文件", type=['xlsx', 'xls'])

# 定义学院排序顺序
COLLEGE_ORDER = [
    "经济与管理学院",
    "法学院",
    "文学与传媒学院", 
    "数据科学与人工智能学院",
    "电子与电气工程学院",
    "机器人工程学院",
    "建筑与能源工程学院",
    "设计艺术学院",
    "外国语学院",
    "创新创业学院"
]

def save_excel(df, output_stream):
    """将DataFrame保存为Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "排序后数据"
    
    # 写入表头
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # 写入数据
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='center')
    
    # 调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_stream)

def find_time_column(df):
    """查找可能的时间列"""
    time_keywords = ['时间', 'date', '开始时间', '结束时间', '开始', '结束', '日期', '备注']
    
    for col in df.columns:
        col_str = str(col).lower()
        for keyword in time_keywords:
            if keyword in col_str:
                return col
    return None

# 主程序
if excel_file is not None:
    try:
        # 读取Excel文件
        df = pd.read_excel(excel_file)
        df.columns = df.columns.str.strip()
        
        # 检查是否找到"学院"列
        if '学院' not in df.columns:
            st.warning("⚠️ 未找到'学院'列，尝试将第二行作为表头...")
            excel_file.seek(0)
            df = pd.read_excel(excel_file, skiprows=1)
            df.columns = df.columns.str.strip()
            
            if '学院' not in df.columns:
                st.error("❌ 无法找到'学院'列。")
                st.write("当前文件中的列名：", df.columns.tolist())
                st.stop()
        
        # 显示原始数据
        st.write(f"总共有 {len(df)} 行数据")
        st.dataframe(df.head())
        
        # 选择排序方式
        sort_method = st.radio("选择排序方式：", ["按学院排序", "按时间列排序"])
        
        if sort_method == "按学院排序":
            # 清理学院列
            df['学院'] = df['学院'].astype(str).str.strip()
            
            # 规范化学院名称
            college_name_mapping = {
                "经管学院": "经济与管理学院",
                "文传学院": "文学与传媒学院",
                "电电学院": "电子与电气工程学院",
                "建工学院": "建筑与能源工程学院",
                "外院": "外国语学院",
                "设艺学院": "设计艺术学院",
                "创业学院": "创新创业学院",
                "数智学院": "数据科学与人工智能学院",
                "电子与电气工程": "电子与电气工程学院",
                "创新与创业学院": "创新创业学院",
                "建筑与能源工程": "建筑与能源工程学院",
                "经管": "经济与管理学院",
                "数据科学与人工智能": "数据科学与人工智能学院",
                "数智":"数据科学与人工智能学院",
            }
            
            def normalize_college_name(name):
                name_clean = str(name).strip()
                return college_name_mapping.get(name_clean, name_clean)
            
            df["学院"] = df["学院"].apply(normalize_college_name)
            
            # 按指定顺序排序
            sorted_dfs = []
            for college in COLLEGE_ORDER:
                college_data = df[df['学院'] == college]
                if not college_data.empty:
                    sorted_dfs.append(college_data)
            
            if sorted_dfs:
                df_sorted = pd.concat(sorted_dfs, ignore_index=True)
                
                # 处理其他学院
                other_colleges = set(df['学院'].unique()) - set(COLLEGE_ORDER)
                if other_colleges:
                    other_data = df[df['学院'].isin(other_colleges)]
                    df_sorted = pd.concat([df_sorted, other_data], ignore_index=True)
                
                st.write("**排序后的数据预览：**")
                st.dataframe(df_sorted.head(10))
                
        else:  # 按时间列排序
            time_column = find_time_column(df)
            
            if time_column:
                st.write(f"使用 '{time_column}' 列进行排序")
                
                # 尝试按时间排序
                try:
                    df['temp_datetime'] = pd.to_datetime(df[time_column], errors='coerce')
                    
                    if df['temp_datetime'].notna().mean() > 0.5:
                        df_sorted = df.sort_values(by='temp_datetime', na_position='last')
                    else:
                        df_sorted = df.sort_values(by=time_column, na_position='last')
                    
                    df_sorted = df_sorted.drop(columns=['temp_datetime'])
                    
                except:
                    df_sorted = df.sort_values(by=time_column, na_position='last')
                
                st.write("**按时间排序后的数据预览：**")
                st.dataframe(df_sorted.head(10))
                
            else:
                st.error("未找到时间列，请确保文件包含时间相关列")
                st.stop()
        
        # 导出文件
        if st.button("📥 导出Excel文件", type="primary"):
            output = io.BytesIO()
            save_excel(df_sorted, output)
            output.seek(0)
            
            filename = "按学院排序.xlsx" if sort_method == "按学院排序" else "按时间排序.xlsx"
            
            st.download_button(
                label=f"下载 {filename}",
                data=output,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    
    except Exception as e:
        st.error(f"处理文件失败: {str(e)}")

else:
    st.info("👆 请先上传Excel文件")




